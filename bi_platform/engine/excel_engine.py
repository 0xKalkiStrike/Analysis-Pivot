"""High-performance Excel / CSV loader based on Polars + fallbacks.

Handles: .xlsx, .xls, .xlsm, .xlsb, .csv, .tsv
Streams chunks for millions of rows and detects worksheets automatically.
"""
from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Iterable

import polars as pl

from ..core.constants import SUPPORTED_ALL, SUPPORTED_CSV, SUPPORTED_EXCEL
from ..core.logger import get_logger
from ..models import Dataset, SheetInfo

log = get_logger(__name__)


class ExcelEngine:
    """Scan folders, discover files, load worksheets."""

    def __init__(self, max_workers: int = 4) -> None:
        self.max_workers = max_workers

    # ------------------------------------------------------------------ discovery
    def scan_folder(self, folder: str | Path, recursive: bool = True) -> list[Path]:
        p = Path(folder)
        if not p.exists():
            return []
        pattern = "**/*" if recursive else "*"
        files = [
            f for f in p.glob(pattern)
            if f.is_file() and f.suffix.lower() in SUPPORTED_ALL
        ]
        log.info(f"Discovered {len(files)} files in {folder}")
        return sorted(files)

    def list_sheets(self, file_path: str | Path) -> list[str]:
        p = Path(file_path)
        suffix = p.suffix.lower()
        if suffix in SUPPORTED_CSV:
            return [p.stem]
        if suffix == ".xlsb":
            import pyxlsb
            with pyxlsb.open_workbook(str(p)) as wb:
                return list(wb.sheets)
        if suffix == ".xls":
            import xlrd
            wb = xlrd.open_workbook(str(p), on_demand=True)
            return wb.sheet_names()
        # xlsx / xlsm
        from openpyxl import load_workbook
<<<<<<< HEAD
        wb = load_workbook(str(p), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
=======
        try:
            wb = load_workbook(str(p), read_only=True, data_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        except Exception:
            return ["Sheet1"]
>>>>>>> a4386bf (Initial commit)

    # ------------------------------------------------------------------ loading
    def load_sheet(
        self,
        file_path: str | Path,
        sheet_name: str | None = None,
        max_rows: int | None = None,
    ) -> Dataset:
        p = Path(file_path)
        suffix = p.suffix.lower()

        if suffix in SUPPORTED_CSV:
            df = self._load_csv(p, max_rows=max_rows)
            sheet = p.stem
        elif suffix in SUPPORTED_EXCEL:
            sheet = sheet_name or self.list_sheets(p)[0]
            df = self._load_excel(p, sheet, max_rows=max_rows)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

        df = self._normalize_headers(df)
        info = SheetInfo(
            file_path=str(p),
            sheet_name=sheet,
            rows=df.height,
            cols=df.width,
            columns=df.columns,
            dtypes={c: str(df.schema[c]) for c in df.columns},
            file_size=p.stat().st_size,
        )
        return Dataset(name=info.display_name, df=df, source=info)

    def load_many(
        self,
        files: Iterable[str | Path],
        sheet_name: str | None = None,
    ) -> list[Dataset]:
        files = list(files)
        datasets: list[Dataset] = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
            futs = {ex.submit(self.load_sheet, f, sheet_name): f for f in files}
            for fut in as_completed(futs):
                try:
                    datasets.append(fut.result())
                except Exception as e:  # pragma: no cover
                    log.error(f"Failed loading {futs[fut]}: {e}")
        return datasets

    # ------------------------------------------------------------------ internals
    def _load_csv(self, p: Path, max_rows: int | None) -> pl.DataFrame:
        sep = "\t" if p.suffix.lower() == ".tsv" else ","
        try:
            return pl.read_csv(
                p, separator=sep, n_rows=max_rows,
                infer_schema_length=1000, try_parse_dates=True,
                ignore_errors=True,
            )
        except Exception:
            return pl.read_csv(
                p, separator=sep, n_rows=max_rows,
                infer_schema_length=0, ignore_errors=True,
            )

    def _load_excel(self, p: Path, sheet: str, max_rows: int | None) -> pl.DataFrame:
        suffix = p.suffix.lower()
        # Polars handles xlsx via fastexcel/openpyxl; use pandas bridge for widest coverage.
        import pandas as pd
        engine_map = {
            ".xlsx": "openpyxl", ".xlsm": "openpyxl",
            ".xls": "xlrd", ".xlsb": "pyxlsb",
        }
        pdf = pd.read_excel(p, sheet_name=sheet, engine=engine_map[suffix], nrows=max_rows)
        # Coerce to strings for object columns with mixed types to avoid Polars schema errors
        for col in pdf.columns:
            if pdf[col].dtype == object:
                pdf[col] = pdf[col].astype("string")
        return pl.from_pandas(pdf)

    def _normalize_headers(self, df: pl.DataFrame) -> pl.DataFrame:
        rename = {}
        seen: dict[str, int] = {}
        for c in df.columns:
            base = str(c).strip()
            if not base or base.lower().startswith("unnamed"):
                base = "column"
            if base in seen:
                seen[base] += 1
                rename[c] = f"{base}_{seen[base]}"
            else:
                seen[base] = 0
                rename[c] = base
        return df.rename(rename)
