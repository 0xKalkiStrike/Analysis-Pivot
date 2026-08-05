"""File and Workspace Discovery Engine.

Recursively scans directories and ZIP archives (including nested ZIPs),
discovers spreadsheets (.xlsx, .xls, .xlsm, .xlsb, .csv, .tsv), detects empty sheets,
corrupted files, password-protected files, and duplicate file hashes.
"""
from __future__ import annotations

import hashlib
import os
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ..core.constants import SUPPORTED_ALL
from ..core.logger import get_logger
from .excel_engine import ExcelEngine

log = get_logger(__name__)


def inspect_sheet_efficiently(file_path: Path, sheet_name: str, excel_engine: ExcelEngine) -> tuple[int, int, list[str]]:
    """Get rows count, cols count and cols names of a sheet extremely efficiently without loading all cell data into memory."""
    suffix = file_path.suffix.lower()
    
    # Handle CSV / TSV
    if suffix in [".csv", ".tsv"]:
        import polars as pl
        sep = "\t" if suffix == ".tsv" else ","
        try:
            df_cols = pl.read_csv(file_path, separator=sep, n_rows=1, ignore_errors=True)
            cols_names = df_cols.columns
            cols_cnt = len(cols_names)
            
            rows_cnt = 0
            with open(file_path, "rb") as f:
                for _ in f:
                    rows_cnt += 1
            rows_cnt = max(0, rows_cnt - 1)
            return rows_cnt, cols_cnt, cols_names
        except Exception:
            pass

    # Handle Excel Spreadsheets
    elif suffix in [".xlsx", ".xlsm"]:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import range_boundaries
            wb = load_workbook(filename=file_path, read_only=True, keep_links=False)
            try:
                ws = wb[sheet_name]
                
                # Use ws.dimensions for instant retrieval without cell parsing
                dims = ws.dimensions
                if dims and ":" in dims:
                    try:
                        _, _, cols_cnt, rows_cnt = range_boundaries(dims)
                    except Exception:
                        rows_cnt = ws.max_row
                        cols_cnt = ws.max_column
                else:
                    rows_cnt = ws.max_row
                    cols_cnt = ws.max_column
                
                if rows_cnt is None or cols_cnt is None:
                    rows_cnt = 0
                    cols_cnt = 0
                    for row in ws.iter_rows(values_only=True):
                        rows_cnt += 1
                        if len(row) > cols_cnt:
                            cols_cnt = len(row)
                
                cols_names = []
                for row in ws.iter_rows(max_row=1, max_col=cols_cnt or 100, values_only=True):
                    cols_names = [str(x) if x is not None else f"Column_{i}" for i, x in enumerate(row)]
                
                final_rows = max(0, rows_cnt - 1) if rows_cnt > 0 else 0
                return final_rows, cols_cnt, cols_names
            finally:
                wb.close()
        except Exception:
            pass
            
    elif suffix == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path, on_demand=True)
            try:
                ws = wb.sheet_by_name(sheet_name)
                rows_cnt = ws.nrows
                cols_cnt = ws.ncols
                cols_names = [str(x) for x in ws.row_values(0)]
                final_rows = max(0, rows_cnt - 1) if rows_cnt > 0 else 0
                return final_rows, cols_cnt, cols_names
            finally:
                wb.release_resources()
        except Exception:
            pass
            
    elif suffix == ".xlsb":
        try:
            from pyxlsb import open_workbook
            with open_workbook(file_path) as wb:
                with wb.get_sheet(sheet_name) as ws:
                    rows_cnt = 0
                    cols_cnt = 0
                    cols_names = []
                    for idx, row in enumerate(ws.iter_rows()):
                        rows_cnt += 1
                        if idx == 0:
                            cols_names = [str(cell.v) if cell.v is not None else f"Column_{i}" for i, cell in enumerate(row)]
                            cols_cnt = len(cols_names)
                    final_rows = max(0, rows_cnt - 1) if rows_cnt > 0 else 0
                    return final_rows, cols_cnt, cols_names
        except Exception:
            pass
            
    try:
        ds = excel_engine.load_sheet(file_path, sheet_name=sheet_name)
        return ds.rows, ds.cols, ds.columns
    except Exception as e:
        log.error(f"Fallback loading failed for {file_path}: {e}")
        return 0, 0, []


@dataclass
class FileDiscoveryInfo:
    file_path: str
    relative_path: str
    file_name: str
    file_size: int
    file_hash: str
    extension: str
    status: str  # "ok" | "corrupted" | "password_protected" | "duplicate" | "unsupported"
    worksheets_count: int = 0
    total_rows: int = 0
    sheets: list[dict[str, Any]] = field(default_factory=list)
    error_message: str | None = None
    is_duplicate: bool = False
    duplicate_of: str | None = None


@dataclass
class DiscoveryReport:
    total_files: int = 0
    total_worksheets: int = 0
    total_records: int = 0
    total_columns: int = 0
    empty_sheets_count: int = 0
    corrupted_files_count: int = 0
    password_protected_files_count: int = 0
    duplicate_files_count: int = 0
    files: list[FileDiscoveryInfo] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_files": self.total_files,
            "total_worksheets": self.total_worksheets,
            "total_records": self.total_records,
            "total_columns": self.total_columns,
            "empty_sheets_count": self.empty_sheets_count,
            "corrupted_files_count": self.corrupted_files_count,
            "password_protected_files_count": self.password_protected_files_count,
            "duplicate_files_count": self.duplicate_files_count,
            "files": [
                {
                    "file_path": f.file_path,
                    "relative_path": f.relative_path,
                    "file_name": f.file_name,
                    "file_size": f.file_size,
                    "file_hash": f.file_hash,
                    "extension": f.extension,
                    "status": f.status,
                    "worksheets_count": f.worksheets_count,
                    "total_rows": f.total_rows,
                    "sheets": f.sheets,
                    "error_message": f.error_message,
                    "is_duplicate": f.is_duplicate,
                    "duplicate_of": f.duplicate_of,
                }
                for f in self.files
            ],
        }


class DiscoveryEngine:
    """Discovers, inspects, and analyzes raw workspace files prior to heavy processing."""

    def __init__(self, excel_engine: ExcelEngine | None = None) -> None:
        self.excel_engine = excel_engine or ExcelEngine()

    @staticmethod
    def _compute_hash(file_path: Path) -> str:
        h = hashlib.sha256()
        with open(file_path, "rb") as f:
            while chunk := f.read(65536):
                h.update(chunk)
        return h.hexdigest()

    def extract_zips_recursively(self, target_dir: str | Path) -> None:
        """Extract any .zip archives found in target_dir recursively into subfolders."""
        target_path = Path(target_dir)
        zip_files = list(target_path.rglob("*.zip"))
        visited = set()

        while zip_files:
            z_path = zip_files.pop()
            if z_path in visited or not z_path.exists():
                continue
            visited.add(z_path)

            dest_folder = z_path.parent / f"_extracted_{z_path.stem}"
            dest_folder.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(z_path, "r") as zf:
                    zf.extractall(dest_folder)
                log.info(f"Extracted {z_path.name} to {dest_folder}")
            except Exception as e:
                log.error(f"Failed to extract {z_path.name}: {e}")

            # Check if new nested zips were unpacked
            new_zips = [p for p in dest_folder.rglob("*.zip") if p not in visited]
            zip_files.extend(new_zips)

    def scan_workspace(self, root_dir: str | Path) -> DiscoveryReport:
        """Recursively scan root_dir for all spreadsheet files and generate a DiscoveryReport."""
        root_path = Path(root_dir).resolve()
        if not root_path.exists():
            return DiscoveryReport()

        # Step 1: Automatically unpack any zip archives present
        self.extract_zips_recursively(root_path)

        # Step 2: Gather all candidate files
        all_candidate_files: list[Path] = []
        for p in sorted(root_path.rglob("*")):
            if p.is_file() and p.suffix.lower() in SUPPORTED_ALL:
                # Skip files larger than 50MB or generated duplicate reports to prevent memory/timeout issues
                if p.stat().st_size > 50 * 1024 * 1024:
                    log.warning(f"Skipping large file during discovery: {p.name}")
                    continue
                if "duplicate_report" in p.name.lower():
                    log.info(f"Skipping duplicate report during discovery: {p.name}")
                    continue
                all_candidate_files.append(p)

        report = DiscoveryReport()
        seen_hashes: dict[str, str] = {}  # hash -> relative_path

        for p in all_candidate_files:
            rel_path = str(p.relative_to(root_path)).replace("\\", "/")
            file_size = p.stat().st_size
            file_hash = self._compute_hash(p)

            info = FileDiscoveryInfo(
                file_path=str(p),
                relative_path=rel_path,
                file_name=p.name,
                file_size=file_size,
                file_hash=file_hash,
                extension=p.suffix.lower(),
                status="ok",
            )

            # Check duplicate file content
            if file_hash in seen_hashes:
                info.is_duplicate = True
                info.duplicate_of = seen_hashes[file_hash]
                info.status = "duplicate"
                report.duplicate_files_count += 1

            seen_hashes[file_hash] = rel_path

            # Attempt inspection of sheets
            try:
                sheets_list = self.excel_engine.list_sheets(p)
                info.worksheets_count = len(sheets_list)
                file_total_rows = 0

                for sheet_name in sheets_list:
                    try:
                        rows_cnt, cols_cnt, cols_names = inspect_sheet_efficiently(p, sheet_name, self.excel_engine)
                        is_empty = (rows_cnt == 0 or cols_cnt == 0)

                        if is_empty:
                            report.empty_sheets_count += 1

                        file_total_rows += rows_cnt
                        report.total_records += rows_cnt
                        report.total_worksheets += 1

                        info.sheets.append({
                            "sheet_name": sheet_name,
                            "rows": rows_cnt,
                            "cols": cols_cnt,
                            "columns": cols_names,
                            "is_empty": is_empty,
                        })
                    except Exception as sheet_err:
                        err_str = str(sheet_err).lower()
                        if "password" in err_str or "protected" in err_str or "encrypted" in err_str:
                            info.status = "password_protected"
                            info.error_message = f"Password protected sheet: {sheet_name}"
                            report.password_protected_files_count += 1
                        else:
                            info.sheets.append({
                                "sheet_name": sheet_name,
                                "rows": 0,
                                "cols": 0,
                                "columns": [],
                                "is_empty": True,
                                "error": str(sheet_err),
                            })

                info.total_rows = file_total_rows

            except Exception as e:
                err_msg = str(e).lower()
                if "password" in err_msg or "protected" in err_msg or "encrypted" in err_msg or "invalidfileexception" in err_msg:
                    info.status = "password_protected"
                    info.error_message = "Password protected or encrypted file"
                    report.password_protected_files_count += 1
                else:
                    info.status = "corrupted"
                    info.error_message = f"Corrupted file: {e}"
                    report.corrupted_files_count += 1

            report.total_files += 1
            report.files.append(info)

        return report
